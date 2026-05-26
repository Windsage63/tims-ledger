from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SHEET_MAPPING_HINTS = {
    "customer": "customers",
    "client": "customers",
    "project": "projects",
    "time": "time_entries",
    "timesheet": "time_entries",
    "expense": "expenses",
    "invoice": "invoices",
    "income": "payments",
    "payment": "payments",
}


def preview_workbook(path: str) -> dict:
    workbook_path = Path(path).expanduser()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook was not found: {workbook_path}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Workbook path must point to an .xlsx or .xlsm file.")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            header_row, headers = _detect_headers(worksheet)
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "header_row": header_row,
                    "headers": headers,
                    "mapping_hint": _mapping_hint(worksheet.title, headers),
                }
            )
        return {"path": str(workbook_path), "sheets": sheets}
    finally:
        workbook.close()


def _detect_headers(worksheet: Any) -> tuple[int | None, list[str]]:
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True),
        start=1,
    ):
        headers = [_clean_header(value) for value in row]
        headers = [header for header in headers if header]
        if len(headers) >= 2:
            return row_number, headers
    return None, []


def _clean_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _mapping_hint(sheet_name: str, headers: list[str]) -> str | None:
    sheet_haystack = sheet_name.lower()
    for keyword, mapping in SHEET_MAPPING_HINTS.items():
        if keyword in sheet_haystack:
            return mapping

    haystack = " ".join(headers).lower()
    for keyword, mapping in SHEET_MAPPING_HINTS.items():
        if keyword in haystack:
            return mapping
    return None
